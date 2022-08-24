from django.shortcuts import render, HttpResponse
from .models import Plates
import openpyxl
import seaborn as sns
import pandas as pd
import matplotlib
import numpy as np
import scipy.optimize as optimization
from matplotlib.ticker import ScalarFormatter
import statistics
from operator import itemgetter
import string
import pickle
from django.core import serializers
from django.conf import settings
from os.path import join, getctime
from os import sep, listdir, mkdir, remove
from copy import deepcopy
import datetime
import shutil
from pathlib import Path
# Make multithreading safe
matplotlib.use('Agg')
import matplotlib.pyplot as plt

version_number = 1.2

def reset_data():
    #set globals
    global totaal; totaal = []
    global check; check = ''
    global check2; check2 = ''
    global dilution; dilution = []
    global seprate_dilution; seprate_dilution = []
    global end_dilution; end_dilution = [] #is this even used anymore? -no, no it's not
    global dictionary; dictionary = {}
    global HD; HD = ''
    global delete; delete = []
    global points_dictionary; points_dictionary = {}
    global mean_ST_dictionary; mean_ST_dictionary = {}
    global mean; mean = 0
    global std; std = 0
    global mean2; mean2 = 0
    global std2; std2 = 0
    global check_cut_off; check_cut_off = 'false'
    global cut_data; cut_data = []
    global outlier_value; outlier_value = 0.0
    global cut_off_value; cut_off_value = 0.0
    global end_result; end_result = {}
    global lower; lower = 0.0
    global upper; upper = 0.0
    global intermediate_dictionary; intermediate_dictionary = {}
    global params_dictionary; params_dictionary = {}
    global final_dictionary; final_dictionary = {}
    global final_list; final_list = []
    global cut_off_value_au; cut_off_value_au = 0
    global unit_name; unit_name = ''
    global row_standard; row_standard = '0'
    global column_standard; column_standard = [0, 0]
    global elisa_type; elisa_type = ''
    global cut_off_type; cut_off_type = ''
    global divide_number; divide_number = 0
    global seprate_dilution; seprate_dilution = []
    global st_finder; st_finder = []
    global dict_st; dict_st = {}
    global list_st_values; list_st_values = []
    global standard; standard = 0
    global rule; rule = 'none'
    global flow; flow = {}
    global last_autosave; last_autosave = datetime.datetime(1970, 1, 1, 0, 0)
    #empty plates from db
    Plates.objects.all().delete()
    #empty pngs from images
    for file in listdir(get_mediapath()):
        if file.endswith('.png'):
            remove(get_mediapath(file))
    return


def get_mediapath(extension=''):
    try:
        mediapath = join(settings.BASE_DIR, 'ELISA_core' + sep + 'static' + sep + 'images' + sep + extension)
    except FileNotFoundError:
        mkdir(join(settings.BASE_DIR, 'ELISA_core' + sep + 'static' + sep + 'images'))
        mediapath = join(settings.BASE_DIR, 'ELISA_core' + sep + 'static' + sep + 'images' + sep + extension)
    return mediapath

def Home(request):
    """
    Input:
        - request: Catches submits from template.
    Output:
        -
    Function:
        - Renders the template Home.html when the page is visited.
    """
    return render(request, 'Home.html', {
            'version': version_number,
        })


def Input_data(request):
    """
    Input:
        - request: Catches submits from template
    Output:
        -
    Function:
        - Checks if the user clicked the button to empty the database and then renders the page with a message
          indicating that it was successfullly emptied. Then checks for if there were any files submitted, if so it will
          send the data to the file_data() function. The variable error is then used to determine if the submitted files
          were incorrectly formatted and shows the corresponding error on the page. If all is ok the page renders with
          a message to inform the user of this. If any other error occurs which is not properly caught, the page will
          still render and inform the user something went wrong.
    """
    try:
        if request.method == 'POST':
            #start pickle magic
            if request.POST.get('download_pickle'):
                session_writeout("Manual Session")
                filename = request.POST.get('Session_name')
                response = HttpResponse(open("Manual Session.ELISA_App", 'rb').read())
                response['Content-Type'] = 'text/plain'
                response['Content-Disposition'] = f'attachment; filename={filename}.ELISA_App'
                return response
            elif request.POST.get('submit_pickle'):
                session_readin(request.FILES['my_pickle'])
                return render(request, 'Input_data.html', {
                    'check': "pickle_upload",
                })
            #end pickle magic
            error = 'correct'
            if request.POST.get('Empty database'):
                reset_data()
                return render(request, 'Input_data.html', {
                    'check': 'correct_emptied',
                })
            if request.POST.get('file_submit'):
                error = file_data(request)
            if error == "file" or error == "extension":
                return render(request, 'Input_data.html', {
                    'check': error,
                })
            else:
                data = Plates.objects.values()
                temp = []
                for i in data:
                    name = i['id'].lower()
                    temp.append(name)
                return render(request, 'Input_data.html', {
                    'check': 'correct', 'files': temp,
                })
        else:
            return render(request, 'Input_data.html')
    except: #todo should really specify this
        return render(request, 'Input_data.html', {
            'check': 'false',
        })


def file_data(request):
    """
    Input:
        - request: Catches submits from template.
    Output:
        - 'file': A string that is used on the Input_data.html to show a specific error message.
        - 'extension': A string that is used on the Input_data.html to show a specific error message.
    Function:
        - If there are no files selected, but the user did click the submit button, a string will be returned to
          catch this error. If the user did submit files they are then checked on proper formatting, if not a string
          will be returned to catch this error. If these two checks are passed the files are then passed on to a
          corresponding function that handles the specific extension.
    """
    if request.FILES.getlist('my_file') == []:
        return "file"
    for file in request.FILES.getlist('my_file'):
        if str(file).split('.')[1] not in ['txt', 'xlsx', 'xls']:
            return 'extension'
    for file in request.FILES.getlist('my_file'):
        if str(file).split('.')[1] == 'txt':
            data = file.readlines()
            data_string = formatting_txt(data, 2)
            database(data_string, file)
        elif str(file).split('.')[1] == 'xlsx':
            data_string = formatting_xlsx(file)
            database(data_string, file)
        elif str(file).split('.')[1] == 'xls':
            data_string = formatting_xls(file)
            database(data_string, file)


def formatting_txt(data, counter):
    """
    Input:
        - data: Nested list with all the rows from a submitted .txt file.
        - counter: A number that is used to differentiate files that need to be decoded.
    Output:
        - data_string: Formatted string with all values seperated by a special character.
    Function:
        - Reads in the data from a nested list and formats it in a way so it can be used in a single long string.
          This string is then returned to the function file_data()
    """
    lines, data_string, formatted_data = list(), "", list()
    for i in data[:-1]:
        if counter == 1:
            lines.append(i.strip())
        elif counter == 2:
            lines.append(i.strip().decode('utf-8'))
    for j in lines:
        line = j.split('\t')
        formatted_data.append(line)
    formatted_data[1].insert(0, '#')
    for i in formatted_data:
        for j in i:
            data_string += j + "="
    return data_string


def formatting_xlsx(file_name):
    """
    Input:
        - file_name: A string that contains the name of a specific file.
    Output:
        - data_string: A formatted string with all values seperated by a special character.
    Function:
        - Opens an excel workbook and reads in all the cells from a specific worksheet.
          It checks if the data is from a spectramax file and if so, sends it to the spectra_data function.
          If not, the data is processed into the data_string and returned to the file_data() function
    """
    wb = openpyxl.load_workbook(file_name)
    active_sheet = wb.active
    excel_data, data_string = list(), ""
    for row in active_sheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    if excel_data[0][0] == "None":
        data_string = spectra_data(excel_data, data_string)
    else:
        del excel_data[1][0]
        del excel_data[0][1:]
        excel_data[1].insert(0, '#')
        for i in excel_data:
            for j in i:
                if j != 'None' and j != ' ':
                    if j == '( + )':
                        data_string += '10' + "="
                    else:
                        data_string += j + "="
    return data_string


def spectra_data(excel_data, data_string):
    """
        Input:
            - excel_data: Nested list with all the rows from a submitted excel file.
            - data_string: An empty string.
        Output:
            - data_string: Formatted string with all values seperated by a special character.
        Function:
            - Reads in the data from a nested list and adds an indentifier column and row.
              The data is converted to data_string and returned to the function file_data().
        """
    for row in range(len(excel_data)):
        del excel_data[row][0]
    title = excel_data[0][0] + " " + excel_data[1][0]
    del excel_data[0][0]
    excel_data[0].insert(0, title)
    top_row = list(range(len(excel_data[0])-1))
    plus_one = [x + 1 for x in top_row]
    top = [str(x) for x in plus_one]
    top.insert(0, "#")
    excel_data.insert(1, top)
    alphabet_list = list(string.ascii_uppercase)
    for row in range(len(excel_data[2:])):
        del excel_data[2:][row][0]
        excel_data[2:][row].insert(0, alphabet_list[row])
    del excel_data[0][1:]
    for i in excel_data:
        for j in i:
            data_string += j + "="
    return data_string


def formatting_xls(file_name):
    """
    Input:
        - file_name: A string that contains the name of a specific file.
    Output:
        - data_string: Formatted string with all values seperated by a special character.
    Function:
        - Creates a dataframe an excel workbook and reads in all the cells from a specific worksheet.
          It create a nested list from this data and processes thid data into the data_string.
          It checks if the data is from a spectramax file and if so, sends it to the spectra_data function.
          If not, the data is processed into the data_string and returned to the file_data() function.
    """
    excel_data, data_string = list(), ""
    df = pd.read_excel(file_name)
    df['Raw Data{Wavelength:415.0}'] = df['Raw Data{Wavelength:415.0}'].fillna('#')
    first = df.columns.values.tolist()
    df_list = df.values.tolist()
    df_list.insert(0, first)
    for row in df_list:
        row_data = list()
        for cell in row:
            if row[0] == "#":
                try:
                    row_data.append(str(int(cell)))
                except:
                    row_data.append(str(cell))
            else:
                row_data.append(str(cell))
        excel_data.append(row_data)
    del excel_data[1][0]
    del excel_data[0][1:]
    excel_data[1].insert(0, '#')
    for i in excel_data:
        for j in i:
            if j != 'None' and j != ' ':
                data_string += j + "="
    return data_string


def database(data_string, file):
    """
    Input:
        - data_string: A formatted string with all values seperated by a special character.
        - file: A string that contains the name of a specific file.
    Output:
        -
    Function:
        - In this function the data_string gets split up in elements by using the special character as seperator.
          These elements are then used to fill in the table in the database.
    """
    split = data_string.split('=')
    plates_instance = Plates.objects.create(
        id=file,
        name=str(split[0]),
        data=data_string
    )


def Plate_layout(request):
    """
    Input:
        - request: Catches submits from template.
        - totaal: An empty list.
        - check: An empty string.
        - row_standard: A string with only one 0
        - column_standard: A list with two zero's
        - elisa_type: An empty string
        - cut_off_type: An empty string
        - unit_name: An empty string
        - standard: An 0
    Output:
        - totaal: A nested list with submitted data from a plate layout file.
        - check: A variable that is used to check if the data is properly read and ready to be formatted into a table.
        - row_standard: A string with as value a number telling which row contains the ST values
        - column_standard: A list with as values numbers telling which columns contain the ST values
        - elisa_type: An string with the numbers 1 or 2.
        - cut_off_type: An string with the numbers 1 or 2.
        - unit_name: An string with it the submitted value
        - standard: An number with it the submitted value
    Function:
        - This function checks if the user has submitted a plate layout file. It first load in the submitted values from the page
          and save them in global variables. After it will check if some of those variables are empty to check if they are equal to None.
          If they are then the variable gets replaced with a zero for column if it was not equal to zero then it gets
          split so the submitted value becomes a list. Then it checks If they submitted a file, if they did not the page will be rendered
          with an error message telling the user to select a file. If the user did select and submit a file it will be
          passed onto the Plate_layout_1() en Plate_layout_2() function. Afterwards it will rerender the page and
          generate a table containing the data from the file. If no file was submitted it will check if the button standart_input
          was pressed. if so then first it wil load in the submitted values from the page and save them in global variables.
          After it will call the function plate_layout3() to get the new standart, then it save some variable that where
           submitted and render the page.
    """
    global check, totaal, row_standard, column_standard, elisa_type, cut_off_type, unit_name, standard, cut_data, flow
    try:
        if request.method == 'POST':
            if request.POST.get('file_submit'):
                elisa_type = request.POST.get('elisa_type')
                cut_off_type = request.POST.get('cut-off_type')
                row_standard = request.POST.get('row_input')
                column_standard = request.POST.get('column_input')
                flow['Select data input type'] = {"1":"Modified/Non-modified ELISA", "2":"General ELISA"}["2"] #todo go for [elisa_type] and replace "2"
                flow['Cut-off or no cut-off'] = {"1":"I want to use HDs to calculate a cut-off", "2":"I don’t want to calculate a cut-off"}[cut_off_type if cut_off_type != None else "1"]
                if row_standard == None:
                    row_standard = 0
                if column_standard == None:
                    column_standard = [0, 0]
                else:
                    column_standard = column_standard.split(',')
            if request.POST.get('file_submit'):
                totaal = []
                if request.FILES.getlist("my_file") == []:
                    check = 'error'
                    return render(request, 'Plate_layout.html', {
                        'check': check, 'totaal': totaal,'row_input': row_standard,
                        'column_input': str(column_standard[0]) + ',' + str(column_standard[1]), 'elisa_type': elisa_type,
                        'cut_off_type': cut_off_type,
                    })
                excel_data = Plate_layout_1(request, "P")
                flow["Plate Layout"] = excel_data #flowline
                totaal = Plate_layout_2(excel_data)
                check = 'go'
                return render(request, 'Plate_layout.html', {
                    'totaal': totaal,
                    'check': check,'row_input': row_standard,
                    'column_input': str(column_standard[0]) + ',' + str(column_standard[1]), 'elisa_type': elisa_type,
                        'cut_off_type': cut_off_type,
                })
            if request.POST.get('standaard_input'):
                cut_data = []
                elisa_type = request.POST.get('elisa_type')
                cut_off_type = request.POST.get('cut-off_type')
                row_standard = request.POST.get('row_input')
                column_standard = request.POST.get('column_input')
                if row_standard == None:
                    row_standard = 0
                if column_standard == None:
                    column_standard = [0, 0]
                else:
                    column_standard = column_standard.split(',')
                Plate_layout_3(request)
                standard = request.POST.get('standaard')
                unit_name = request.POST.get('unit')
                flow['ST values of all plates'] = request.POST.get('standaard')
                flow['Divide number'] = request.POST.get('divide')
                check = 'go'
                return render(request, 'Plate_layout.html', {
                    'totaal': totaal, 'check': check, 'row_input': row_standard,
                    'column_input': str(column_standard[0]) + ',' + str(column_standard[1]),
                    'standard': standard, 'divide': divide_number, 'unit': unit_name, 'elisa_type': elisa_type,
                        'cut_off_type': cut_off_type,
                })
        else:
            return render(request, 'Plate_layout.html', {
                'totaal': totaal, 'check': check, 'row_input': row_standard,
                'column_input': str(column_standard[0]) + ',' + str(column_standard[1]),
                'standard': standard, 'divide': divide_number, 'unit': unit_name, 'elisa_type': elisa_type,
                        'cut_off_type': cut_off_type,
            })
    except:
        row_standard = 0
        column_standard = [0, 0]
        return render(request, 'Error.html', {
            'error': 'There was an incorrect input. Please return to the Plate Layout page and try again.',
        })

def Plate_layout_1(request, check_type):
    """
    Input:
        - request: Catches submits from template.
    Output:
        - excel_data: A nested list with the data from a plate layout file.
    Function:
        - This function reads the submitted file and converts it to a nested list with all the data from that specific
          file. This nested list is then returned to the Plate_layout() function.
    """
    if check_type == 'P':
        excel_file = request.FILES["my_file"]
    elif check_type == 'D':
        excel_file = request.FILES["dilution_file"]
    elif check_type == 'D2':
        excel_file = request.FILES["dilution_file2"]
    wb = openpyxl.load_workbook(excel_file)
    active_sheet = wb.active
    excel_data = list()
    for row in active_sheet.iter_rows():
        row_data = list()
        for cell in row:
            if type(cell.value) == float: #todo look at this, is this needed?
                row_data.append(str(round(cell.value)))
            else:
                row_data.append(str(cell.value))
        excel_data.append(row_data)
    return excel_data


def Plate_layout_2(excel_data):
    """
    Input:
        - excel_data: A nested list with the data from a plate layout file.
    Output:
        - totaal: A nested list with the data from a plate layout file that is properly formatted and stripped of
          None's.
    Function:
        - This function reads every line in the nested list excel_data, determines the max rows per plate
          and deletes all the None's then inserts values so the lists have the same length.
          Finally it appends the formatted lists to the nested totaal list and
          returns this nested list to the Plate_layout() function.
    """
    temp, counter = [], 0
    length_empty = 0
    tot_rows = len(excel_data)
    for x in range(len(excel_data)):
        if x != 0:
            if 'late' in excel_data[x][0]:
                rows = x-1
                break
            else:
                rows = tot_rows
    for i in excel_data: #row
        k = [e for e in i if e != ('None')] #per row append e if e isn't none
        if length_empty != 0 and len(k) != 0:
            for g in range(length_empty):
                if k[0].isalpha():
                    if i[g] == 'None':
                        k.insert(g, 'Empty')
        if len(k) != 0:
            if counter == 1:
                k.insert(0, '#')
                length_empty = len(k)
            temp.append(k)
            counter += 1
            if counter == (rows):
                totaal.append(temp)
                counter = 0
                temp = []
    return totaal


def Plate_layout_3(request):
    """
    Input:
        - request: Catches submits from template.
    Output:
        - divide_number: in this varibale comes a divide number the user submitted
        - dict_st:
    Function:
        - This function retrieves the submitted ST value the user inputted. This value is then used and divided by two
          for every row in the plate layout file. The last row get a # as value since these values are supposed to be
          zero. When clicking the submit button the page gets reloaded and the table gets filled, so there is no return.
    """
    global divide_number, st_finder, dict_st, list_st_values
    values = request.POST.get('standaard')
    divide_number = request.POST.get('divide')
    list_st_str = []
    list_st_int = []
    list_divide = []
    dict_st = {}
    for index, i in enumerate(totaal): #total is a global, function works cause total has a nested list which is mutable cause of python magic
        for j in range(len(i)):
            list_divide.append(values)
            list_st_str.append('st_' + str(j+1))
            list_st_int.append(round(float(values), 3))
            if 'st_' + str(j+1) not in dict_st and len(list_st_values) == 0:
                dict_st['st_' + str(j+1)] = []
            elif len(list_st_values) != 0:
                if len(dict_st) <= 7:
                    if list_st_values[j] not in dict_st:
                        dict_st[str(list_st_values[j])] = []
            values = float(values) / float(divide_number)
        for j in range(len(i)):
            for k in range(len(i[j])):
                for d in range(len(i)):
                    if str(i[j][k]).lower() == list_st_str[d]:
                        dict_st[i[j][k].lower()].append([index, j, k])
                        if index == 0 and str(i[j][k]).lower() == 'st_1':
                            st_finder = [0, j, k]
                        i[j][k] = round(float(list_divide[d]), 3)
                    elif str(i[j][k]).lower() == 'blanco':
                        i[j][k] = "#"
                    elif len(list_st_values) != 0:
                        if d <= 7:
                            if str(i[j][k]) == str(list_st_values[d]):
                                dict_st[str(i[j][k])].append([index, j, k])
                                i[j][k] = round(float(list_divide[d]), 3)
    list_st_values = deepcopy(list_st_int[:8])


def Dilutions(request):
    """
    Input:
        - request: Catches submit from template.
        - seprate_dilution: An empty list
    Output:
        - dilution: A nested list with the dilution of the plate. The values are all numbers with as type string.
        - seprate_dilution: A nested list with the names of plates
    Function:
        - The function checks if a file is submitted, if so then it gives it to other functions in order to read it.
          Then it changes a function to 'go' to allow it to be shown on the website. Then it checks of dilution only has 1
          plate or more and if its more is it the same size as plate layout.
          the function also checks if the user submitted the options so combine plate names with one dilution file. then
          it checks if the user submitted another file inorder to add it to the dilution variable.
    """
    global dilution, check2, seprate_dilution
    show = 'no'
    if request.method == 'POST':
        if request.POST.get('file_submit'):
            if request.FILES.getlist("dilution_file") == []:
                check2 = 'error'
                return render(request, 'Dilutions.html', {
                    'check': check2, 'dilution': dilution,
                    'show' : show,
                })
            dilution = []
            seprate_dilution = []
            excel_data = Plate_layout_1(request, 'D')
            dilution = Dilutions_1(excel_data)
            check2 = 'go'
            if len(dilution) != len(totaal) and len(dilution) != 1:
                check2 = 'nope'
            return render(request, 'Dilutions.html', {
                'dilution': dilution,
                'check': check2,
                'show': show,
            })
        if request.POST.get('plate_belong1'):
            show = 'yes'
            dilution_1 = request.POST.get('dilution_1')
            dil_list = list(dilution_1.split(", "))
            seprate_dilution.append(dil_list)
            return render(request, 'Dilutions.html', {
                'dilution': dilution,
                'check': check2,
                'dilution_v1' : dilution_1,
                'show': show, })
        if request.POST.get('file_submit2'):
            if request.FILES.getlist("dilution_file2") == []:
                check2 = 'error'
                return render(request, 'Dilutions.html', {
                    'check': check2, 'dilution': dilution,
                    'show' : show,
                })
            excel_data = Plate_layout_1(request, 'D2')
            dilution = Dilutions_1(excel_data)
            check2 = 'go'
            dilution_2 = request.POST.get('dilution_2')
            dil_list = list(dilution_2.split(", "))
            seprate_dilution.append(dil_list)
            return render(request, 'Dilutions.html', {
                'dilution': dilution,
                'check': check2,
                'show': show,
            })
    else:
        return render(request, 'Dilutions.html', {
            'dilution': dilution, 'check': check2,
            'show' : show,})


def Dilutions_1(excel_data):
    global dilution
    """
    Input:
        - excel_data: A nested list with the data from a plate layout file.
    Output:
        - dilution: A nested list with the data from a dilution file that is properly formatted and stripped of
          None's.
    Function:
        - This function reads every line in the nested list excel_data, determines the max rows per plate
          and deletes all the None's then inserts values so the lists have the same length.
          Finally it appends the formatted lists to the nested totaal list and
          returns this nested list to the Dilution() function.
    """
    temp, counter = [], 0
    length_empty = 0
    tot_rows = len(excel_data)
    for x in range(len(excel_data)):
        if x != 0:
            if 'late ' in excel_data[x][0]:
                rows = x - 1
                break
            else:
                rows = tot_rows
    for i in excel_data:
        k = [e for e in i if e != ('None')]
        if length_empty != 0 and len(k) != 0:
            for g in range(length_empty):
                if k[0].isalpha():
                    if i[g] == 'None':
                        k.insert(g, 'Empty')
        if len(k) != 0:
            if counter == 1:
                k.insert(0, '#')
                length_empty = len(k)
            temp.append(k)
            counter += 1
            if counter == rows:
                dilution.append(temp)
                counter = 0
                temp = []
    return dilution


def Visualize_data(request):
    """
    Input:
        - request: Catches submits from template.
        - dictionary: An empty dictionary.
        - HD: An empty string.
        - delete: An empty list.
        - points_dictionary: An empty dictionary.
        - totaal: A nested list with the data from a plate layout file that is properly formatted and stripped of
          None's.
    Output:
        - dictionary: A dictionary with as key the plate names and the value a nested list with in it the OD and RBG
          code.
        - delete: A list with selected plate names which are not allowed in end results.
        - HD: A string with the name of the healthy donor plate.
        - points_dictionary: A dictionary with as key the name of the plate and as value a list of the chosen lower and
          upper points of that plate.
    Function:
        - In this function a dictionary is created to save the chosen lower and upper points of the linear part in
          specific graph. This is saved in the dictionary points_dictionary. Then all the data from the database is
          loaded in and the background noise is calculated and deducted from every single cell. Then a RGB code is
          calculated to use in the table on the page. All the calculated variables are then stored in the 'dictionary'.
          if the global totaal is not empty the create_graph() function is called. The page is then rendered and will
          generate tables, graphs, checkboxes and dropdowns. If any of the code raises and error it is caught with the
          except statement. This will then render an error on the page itself without any of the tables or graphs.
    """
    try:
        global dictionary
        global HD
        global delete
        if request.method == 'POST':
            if request.POST.get("Confirm1"):
                HD = request.POST['HD']
            elif request.POST.get("Confirm2"):
                HD = 'None'
            bottom = request.POST.getlist('top')
            top = request.POST.getlist('bottom')
            counter = 0
            for keys in dictionary:
                points_dictionary[keys] = [top[counter], bottom[counter]]
                counter += 1
            delete = request.POST.getlist('delete')
        data = Plates.objects.values()
        counter = 0
        nested = []
        temp = []
        for i in data:
            name = i['id'].lower()
            lines = i['data'].split('=')[:-1]
            number1 = lines[106].replace(',', '.')
            number2 = lines[107].replace(',', '.')
            calculation = ((float(number1) + float(number2))/2)
            mean = round(calculation, 3)
            max = 0.0
            new_lines = []
            position = lines.index("A")
            for k in lines[:position]:
                new_lines.append(k)
            for index, x in enumerate(lines[position:]):
                if x.isdigit():
                    x = str(float(x))
                new_lines.append(x)
            for x in new_lines[position:]:
                if x[0].isdigit():
                    if float(x) > max:
                        max = float(x)
            for j in new_lines[1:]:
                if ',' in j:
                    new = float(j.replace(',', '.')) - mean
                    c_color = max - new
                    times = 255/max
                    color = c_color*times
                    DCO = round(new, 3)
                    temp.append([DCO, (color, 255, color)])
                elif '.' in j:
                    new = float(j) - mean
                    c_color = max - new
                    times = 255/max
                    color = c_color*times
                    DCO = round(new, 3)
                    temp.append([DCO, (color, 255, color)])
                else:
                    temp.append([j, (255, 255, 255)])
                counter += 1
                if counter == 13:
                    nested.append(temp)
                    counter = 0
                    temp = []
            dictionary[name] = nested
            nested = []
        if totaal != []:
            create_graph(dictionary)
        return render(request, 'Visualize_data.html', {
            'dictionary': dictionary,
            'cut_off_type': cut_off_type,
        })
    except:
         return render(request, 'Error.html', {
             'error': 'An error occurred, please be sure to load in the plate layout file and choose a ST value on the '
                      'Plate Layout page.',
         })


def create_graph(dictionary):
    """
    Input:
        - dictionary: A dictionary with as key the plate names and the value a nested list with in it the OD and RBG
          code.
    Output:
        - mean_ST_dictionary: A dictionary with as key the names of the plates and the value the average score of ODs.
    Function:
        - This function manages the proper formatting of all the data that is needed to create a graph per plate. This
          is done by creating the global dictionary mean_ST_dictionary. The name of the plate is used as key and the
          value consists of a list calculated average ST values of each row. This data is then used to create each
          graph.
    """
    global mean_ST_dictionary
    conc = totaal[st_finder[0]][st_finder[1]][st_finder[2]]
    x_list = [conc]
    number_loop = len(totaal[0]) - 4
    for i in range(number_loop):
        conc = float(conc)/int(divide_number)
        x_list.append(conc)
    y_list = []
    temp = []
    count_plate = 0
    for values in dictionary.values():
        for elements in dict_st.values():
            if len(elements) != 0:
                pos1 = elements[count_plate][1]
                pos2 = elements[count_plate][2]
                pos3 = elements[count_plate+1][1]
                pos4 = elements[count_plate+1][2]
                mean = ((float(values[pos1-1][pos2][0]) + float(values[pos3-1][pos4][0]))/2)
                temp.append(round(mean, 3))
        count_plate += 2
        y_list.append(temp)
        temp = []
    counter = 0
    for keys in dictionary:
        mean_ST_dictionary[keys] = y_list[counter]
        counter += 1
    counter = 0
    for key in dictionary:
        guess = [1, 1, 1, 1, 1]
        params, params_coveriance = optimization.curve_fit(formula, x_list, y_list[counter], guess, maxfev=5000)
        intermediate_list(key, params)
        x_min, x_max = np.amin(x_list), np.amax(x_list)
        xs = np.linspace(x_min, x_max, 1000)
        plt.scatter(x_list, y_list[counter])
        plt.plot(xs, formula(xs, *params))
        plt.xscale('log')
        plt.grid()
        ax = plt.gca()
        plt.xticks([1.0, 10, 100])
        ax.xaxis.set_major_formatter(ScalarFormatter())
        plt.savefig(get_mediapath(str(key) + ".png"))
        plt.close()
        counter += 1


def formula(x, A, B, C, D, E):
    """
    Input:
        - x: A given x-value.
        - A: Value of minimal asymptote.
        - B: Value of steepness.
        - C: Value of inflection point.
        - D: Value of maximal asymptote.
        - E: Value of asymmetry factor.
    Output:
        - return: A y-value for a given x-value.
    Function:
        - This function returns the y-value (OD) of a given x-value (concentration).
    """
    return D + (A-D)/(np.power((1 + np.power((x/C), B)), E))


def Cut_off(request):
    """
    Input:
        - request: Catches submits form template.
        - mean: Has a zero number.
        - std: Has a zero number.
        - mean2: Has a zero number.
        - std2: Has a zero number.
        - cut_data: An empty list
        - row_standard: A string with as value a number telling which row contains the ST values
        - column_standard: A list with as values numbers telling which columns contain the ST values
        - check_cut_off: The string false
        - outlier_value: A float with zero
        - cut_off_value: A float with zero
        - cut_off_type: A string with the number 1 or 2
        - elisa_type: A string with the number 1 or 2
        - dictionary: A dictionary with as key the plate names and the value a nested list with in it the OD and RBG code.
        - HD: An string with the name of the selected healthy donor plate.
    Output:
        - request: Catches submits form template.
        - mean: Here comes the average score for the outlier.
        - std: Here comes the std score for the outlier
        - mean2: Here comes the average score for the cut-off.
        - std2: Here comes the std score for the cut-off
        - cut_data: Here comes a list with the OD scores from the healthy donor plate
        - check_cut_off: A string with True
        - outlier_value: A value with the outlier score
        - cut_off_value: A value with the cut-off value
    Function:
        - This function starts by looking if cut_off_type is equal to 2, if so that means the user did,t want a cutt off
         so an error page is render with the text that the user did not want a cut off. then it checks if a button is
         pressed if not than it will create the first swarm plot for the outliers. This is done by looking which type of
         analysis this is, is it type 1 then it contains mod/non-mods so the data needs to be filtered for it. in order to
         do that it needs to look at the row_standard to see if the variable is equal to zero. if so than that means the ST values
         are in the columns, if not then they are in the rows. if the row_standard is zero the function will check if the
         value is not the same as the first or second position in column_standard. Because if they are the same then
         that means that value is the ST value. count_the_mod wil check if the values are mod values if count_the_mod is
         bigger than 5 that means the values are now the non-mod values. For the rows it checks if values is not the same
         as row_standard and then checks if value is bigger than mod_lengths, if so then that means that value is a non-mod
         value and isn't allowed in the data. all of this is the same for when elisa_type is equal to 2 but without the
         count_the_mod, mod_length parts.
         When a button is pressed it will look into which button was pressed. IF the outlier_submit button was pressed.
         Then it will create the swarm plot for te cut-ff. If the button from cut_off_submit was pressed a cut-off is
         calculated.
    """
    try:
        global mean
        global std
        global mean2
        global std2
        global cut_data
        global check_cut_off
        global outlier_value
        global cut_off_value
        global elisa_type
        cut_dict = {}
        if cut_off_type == '2':
            return render(request, 'Error.html', {
                'error': 'The option to not use a cut-off was selected. Move on to intermediate result to continue'
                         ' the application.'
            })
        if request.method == 'POST':
            if request.POST.get('outlier_submit'):
                input1 = request.POST.get('input1')
                input2 = request.POST.get('input2')
                outlier_value = (float(input1) * mean) + (float(input2) * std)
                outlier_value = round(outlier_value, 3) #TODO Flow for formula and outlier
                new_y_list = []
                for data in cut_data:
                    if data < outlier_value:
                        new_y_list.append(data)
                cut_dict['New_OD'] = new_y_list
                mean2 = round(statistics.mean(new_y_list), 3)
                std2 = round(statistics.stdev(new_y_list), 3)
                df = pd.DataFrame(data=cut_dict)
                ax = sns.swarmplot(data=df, y="New_OD")
                ax = sns.boxplot(data=df, y="New_OD", color='white')
                plt.savefig(get_mediapath('swarmplot2.png'))
                plt.close()
                check_cut_off = 'true'
                return render(request, 'Cut_off.html', {
                    'mean': mean,
                    'std': std,
                    'mean2': mean2,
                    'std2': std2,
                    'check': check_cut_off,
                    'outlier_value': outlier_value,
                    'cut_off_value': cut_off_value,
                })
            elif request.POST.get('cut_off_submit'):
                input3 = request.POST.get('input3')
                input4 = request.POST.get('input4')
                cut_off_value = (float(input3) * mean2) + (float(input4) * std2)
                cut_off_value = round(cut_off_value, 3) #TODO Flow for formula and cut-off
                return render(request, 'Cut_off.html', {
                    'mean': mean,
                    'std': std,
                    'mean2': mean2,
                    'std2': std2,
                    'check': check_cut_off,
                    'outlier_value': outlier_value,
                    'cut_off_value': cut_off_value,
                })
        elif cut_data == []:
            if elisa_type == '1':
                for i, j in dictionary.items():
                    if HD == i:
                        for values in range(len(j)):
                            count_the_mod = 0
                            for value in range(len(j[values])):
                                if type(j[values][value][0]) != str:
                                    if int(row_standard) == 0:
                                        if value != int(column_standard[0]):
                                            if value != int(column_standard[1]):
                                                if count_the_mod < 5:
                                                    cut_data.append(j[values][value][0])
                                                    count_the_mod += 1
                                    elif int(row_standard) != values:
                                        mod_length = len(j[values])/2
                                        if value <= mod_length:
                                            cut_data.append(j[values][value][0])
            elif elisa_type == '2':
                for i, j in dictionary.items():
                    if HD == i:
                        for values in range(len(j)):
                            for value in range(len(j[values])):
                                if type(j[values][value][0]) != str:
                                    if int(row_standard) == 0:
                                        if value != int(column_standard[0]):
                                            if value != int(column_standard[1]):
                                                cut_data.append(j[values][value][0])
                                    elif int(row_standard) != values:
                                        cut_data.append(j[values][value][0])
            cut_dict["OD"] = cut_data
            mean = round(statistics.mean(cut_data), 3)
            std = round(statistics.stdev(cut_data), 3)
            df = pd.DataFrame(data=cut_dict)
            ax = sns.swarmplot(data=df, y="OD")
            ax = sns.boxplot(data=df, y="OD", color='white')
            plt.savefig(get_mediapath('swarmplot.png'))
            plt.close()
            return render(request, 'Cut_off.html', {
                'mean': mean,
                'std': std,
                'check': check_cut_off,
                'outlier_value': outlier_value,
                'cut_off_value': cut_off_value,
            })
        return render(request, 'Cut_off.html', {
            'mean': mean,
            'std': std,
            'mean2': mean2,
            'std2': std2,
            'check': check_cut_off,
            'outlier_value': outlier_value,
            'cut_off_value': cut_off_value,
        })
    except:
        return render(request, 'Error.html', {
            'error': 'An error occurred, please be sure to select the plate with the healthy donor data on the '
                     'Visualize data page.'
        })


def formula2(y, A, B, C, D, E):
    """
    Input:
        - y: A given y-value.
        - A: Value of minimal asymptote.
        - B: Value of steepness.
        - C: Value of inflection point.
        - D: Value of maximal asymptote.
        - E: Value of asymmetry factor.
    Output:
        - return: A x-value for a given y-value.
    Function:
        - This function returns the x-value (concentration) of a given y-value (OD).
    """
    return C*(np.power((np.power(((A-D)/(-D+float(y))), (1/E))-1), (1/B)))


def Intermediate_result(request):
    """
    Input:
        - request: Catches submits form template.
        - end_result: An empty dictionary.
        - lower: The number zero.
        - upper: The number zero.
        - elisa_type: A string with the number 1 or 2
        - HD: An string with the name of the selected healthy donor plate.
        - points_dictionary: A dictionary with as key the name of the plate and as value a list of the chosen lower and
                            upper points of that plate.
        - intermediate_dictionary: A dictionary with as keys the plates names and the value a nested list. With in it
          a sample ID and the calculated au/ml.
        - delete: A list with selected plate names which are not allowed in end results.
        - mean_ST_dictionary: A dictionary with as key the names of the plates and the value the average score of ODs.
        - dilution: A nested list with the dilution of the plates. the values are all numbers with as type string.
        - seprate_dilution: a nested list with the names of plates
    Output:
        - end_result: A dictionary with as key the name of the plates and the values a nested list. the nested list
                      have as values first a sample id, second the au/ml, and third a 1 or 2 or 3.
        - lower: An au/ml score from the lowest chosen value.
        - upper: An au/ml score from the highest chosen value.
    Function:
        - The function first checks which plates it should not to be looking if the plate names in intermediate_dictionary
          are in delete. Then it looks if seprate_dilution is empty or not after this check top and bot will be filled
          with the dilution belonging to that plate. After that it will look if elisa_type is equal to 2 or smaller
          then count_mod 5. this is done to filter out the non-mod values if needed.
          if it passes the if-statements it will look if the values from end_result are smaller or bigger
          then lower or upper. If the value is smaller
          than lower it gets a 1, if higher than upper it gets a 3. If nether than it gets a 2. When the list is filled
          it gets sorted by sample ID and everything gets put into one list. also two other list are made one with the
          last 20 of the below values and the first 20 of the linear values. and the other list with the 20 last of the
          linear values and the 20 first of the above values
    """
    try:
        global end_result
        global lower
        global upper
        global end_result
        end_result = {}
        kopie_dict = deepcopy(intermediate_dictionary)
        for key, values in kopie_dict.items():
            if key not in delete:
                end_result[key] = values
        temp0 = []
        temp1 = []
        temp2 = []
        temp3 = []
        temp4 = []
        mod_check_colum1 = int(column_standard[0]) - 1
        mod_check_colum2 = int(column_standard[1]) - 1
        row_check = 0
        for key, values in end_result.items():
            mean_ST_dictionary[key].reverse()
            top = mean_ST_dictionary[key][int(points_dictionary[key][1]) - 1]
            bot = mean_ST_dictionary[key][int(points_dictionary[key][0]) - 1]
            if len(seprate_dilution) != 0:
                for sep in seprate_dilution[0]:
                    if sep == key[key.index("plate"):key.index("plate") + 7] or sep == key[key.index("plate"):key.index("plate") + 8] or sep == key[key.index("plate"):key.index("plate") + 9]:
                        string_top = formula2(top, *params_dictionary[key]) * int(dilution[0][3][3])
                        string_bot = formula2(bot, *params_dictionary[key]) * int(dilution[0][3][3])
                for sep in seprate_dilution[1]:
                    if sep == key[key.index("plate"):key.index("plate") + 7] or sep == key[key.index("plate"):key.index("plate") + 8] or sep == key[key.index("plate"):key.index("plate") + 9]:
                        string_top = formula2(top, *params_dictionary[key]) * int(dilution[1][3][3])
                        string_bot = formula2(bot, *params_dictionary[key]) * int(dilution[1][3][3])
            elif len(dilution) == 1:
                string_top = formula2(top, *params_dictionary[key]) * int(dilution[0][3][3])
                string_bot = formula2(bot, *params_dictionary[key]) * int(dilution[0][3][3])
            else:
                for d in range(len(dilution)):
                    if dilution[d][0][0].lower() == key[key.index("plate"):key.index("plate") + 7] or dilution[d][0][0].lower() == key[key.index("plate"):key.index("plate") + 8] or dilution[d][0][0].lower() == key[key.index("plate"):key.index("plate") + 9]:
                        string_top = formula2(top, *params_dictionary[key]) * int(dilution[d][3][3])
                        string_bot = formula2(bot, *params_dictionary[key]) * int(dilution[d][3][3])
            count_mod = 0
            count_mod2 = 0
            row_check += 1
            for value in values:
                if row_check != int(row_standard):
                    if count_mod == 12:
                       count_mod = 0
                       count_mod2 = 0
                    if elisa_type == '2':
                        count_mod2 = 0
                    if count_mod != int(mod_check_colum1) and count_mod != int(mod_check_colum2) and count_mod2 < 5 or int(row_standard) != 0 and count_mod2 < 5:
                        count_mod2 += 1
                        if type(value[1]) == str:
                            if len(value) == 2:
                                if float(value[1]) < bot:
                                    temp0.append([value[0]] + ['<' + str(round(string_bot, 3))] + ["below"])
                                else:
                                    temp4.append([value[0]] + ['>' + str(round(string_top, 3))] + ['linear'])
                            else:
                                value[2] = 1
                                temp0.append(value[:3])
                        elif int(value[1]) <= float(string_bot):
                            if len(value) == 2:
                                temp1.append(value + ['below'])
                            else:
                                value[2] = 1
                                temp1.append(value[:3])
                        elif int(value[1]) >= float(string_top):
                            if len(value) == 2:
                                temp3.append(value + ['above'])
                            else:
                                value[2] = 3
                                temp3.append(value[:3])
                        else:
                            if len(value) == 2:
                                temp2.append(value + ['linear'])
                            else:
                                value[2] = 2
                                temp2.append(value[:3])
                count_mod += 1
            mean_ST_dictionary[key].reverse()
        sorted_temp1 = sorted(temp1, key=itemgetter(1))
        sorted_temp2 = sorted(temp2, key=itemgetter(1))
        sorted_temp3 = sorted(temp3, key=itemgetter(1))
        complete_list = temp0 + sorted_temp1 + sorted_temp2 + sorted_temp3 + temp4
        if len(sorted_temp1) < 20:
            low_list = sorted_temp1 + sorted_temp2[:20]
        else:
            pos = len(sorted_temp1) - 20
            low_list = sorted_temp1[pos:] + sorted_temp2[:20]
        if len(sorted_temp2) < 20:
            up_list = sorted_temp2 + sorted_temp3[:20]
        else:
            pos = len(sorted_temp2) - 20
            up_list = sorted_temp2[pos:] + sorted_temp3[:20]
        if request.method == 'POST':
            if request.POST.get('limit_submit_l'):
                lower = request.POST.get('lower')
                return render(request, 'Intermediate_result.html', {
                    'complete_list': complete_list,
                    'unit': unit_name,
                    'lower': lower,
                    'upper': upper,
                    'limit_list': up_list,
                    'check': 'go_up'
                })
            if request.POST.get('limit_submit_u'):
                upper = request.POST.get('upper')
        return render(request, 'Intermediate_result.html', {
            'complete_list': complete_list,
            'unit': unit_name,
            'lower': lower,
            'upper': upper,
            'limit_list' : low_list,
            'check' : 'go_low'
        })
    except:
        return render(request, 'Error.html', {
            'error': 'An error occurred, please make sure you have selected the healthy donor plate and confirming '
                     'your preferences on the visualize Data page.'
        })


def intermediate_list(key, params):
    """
    Input:
        - key: A string with the name of the plate.
        - params: The optimal values for the curve fit. A: minimal asymptote, B: steepness, C: inflection point
                 D: maximal asymptote, E: Asymmetry factor.
        - intermediate_dictionary: An empty dictionary.
        - params_dictionary: An Empty dictionary.
        - totaal: Nested list with submitted data from a plate_layout file.
        - end_dilution: A nested list with the dilution of the plates. the values are all numbers with as type string.
    Output:
        - intermediate_dictionary: A dictionary with as keys the plates names and the value a nested list. With in it
          a sample ID and the calculated au/ml.
        - params_dictionary: In this dictionary the keys are the name of the plates and the values the params.
    Function:
            - First the function takes a look which plate from visualize data belongs with the plate from
              plate-layout by comparing the number in de plate name. then it save the params in the params_dictionary
              and calls the formula2 function, so the result can be multiplied by the dilution score to get the au/ml.
    """
    global intermediate_dictionary
    for options in range(len(totaal)):
        if totaal[options][0][0].lower() == key[key.index("plate"):key.index("plate") + 7] or totaal[options][0][0].lower() == key[key.index("plate"):key.index("plate") + 8] or totaal[options][0][0].lower() == key[key.index("plate"):key.index("plate") + 9]:
            position = options
    list1 = []
    count_plate = 0
    for i, j in dictionary.items():
        if i == key:
            count_plate += 2
            params_dictionary[key] = params
            for values in range(len(j)):
                if values != 0:
                    for value in range(len(j[values])):
                        if value != 0:
                            result = formula2(j[values][value][0], *params)
                            if np.isnan(result):
                                result = str(j[values][value][0])
                            else:
                                if len(seprate_dilution) == 0:
                                    if len(dilution) == 1:
                                        result *= int(dilution[0][values][value])
                                    else:
                                        for dil in range(len(dilution)):
                                            if dilution[dil][0][0].lower() == key[key.index("plate"):key.index("plate") + 7] or dilution[dil][0][0].lower() == key[key.index("plate"):key.index("plate") + 8] or dilution[dil][0][0].lower() == key[key.index("plate"):key.index("plate") + 9]:
                                                result *= int(dilution[dil][values][value])
                                else:
                                    for d in range(len(seprate_dilution)):
                                        for g in seprate_dilution[d]:
                                            if g == key[key.index("plate"):key.index("plate") + 7] or g == key[key.index("plate"):key.index("plate") + 8] or g == key[key.index("plate"):key.index("plate") + 9]:
                                                result *= int(dilution[d][values][value])

                                result = round(result, 3)
                            list1.append([totaal[position][values + 1][value], result])
                intermediate_dictionary[i] = list1



def End_results(request):
    """
    Input:
        - request: Catches submits form template.
        - HD: An string with the name of the selected healthy donor plate.
        - final_dictionary: An empty dictionarty list.
        - final_list: An empty list.
        - cut_off_value_au: The number zero.
        - rule: string with none
        - row_standard: A string with as value a number telling which row contains the ST values
        - column_standard: A list with as values numbers telling which columns contain the ST values
        - elisa_type: A string with the number 1 or 2
        - cut_off_type: A string with the number 1 or 2
        - dictionary: A dictionary with as key the plate names and the value a nested list with in it the OD and RBG code.
        - cut_off_value: It has contains a float which is the OD given by the cut_off function
        - params_dictionary: In this dictionary the keys are the name of the plates and the values the params.
        - end_dilution: A nested list with the dilution of the plates. the values are all numbers with as type string.
        - end_result: A dictionary with as key the name of the plates and the values a nested list. the nested list
                      have as values first a sample id, second the au/ml, and third a 1 or 2 or 3.
        - lower: An au/ml score.
        - delete: A list with selected plate names which are not allowed in end results.
    Output:
        - final_dictionary: The dictionary is now filled and has as key sampleID which start with 1 and goes up by 1
                            with every now result. the values are a list with as first a sample id, second an 1 or 2,
                            and third the au/ml.
        - final_list: A list with as first a sample id, second an 0 or 1 , third the au/ml, fourth an OD and if requested
                      an non-mod OD.
        - cut_off_value_au: An au/ml, calculated from the OD from cut_off_value and the params from params_dictionary.
        - rule: string with 1 or 2 or 3 or 4.
        - end_result: A dictionary with as key the name of the plates and the values a nested list. the nested list
                      have as values first a sample id, second the au/ml, and third a 1 or 2 or 3.
    Function:
        - The function first checks if any button was pressed, if there were any buttons pressed it then check which.
          first it will check if the user wanted a cutt_off by checking cutt_off type.
          second it will add some additional information to end_result dictionary. But in order to that it needed to see
          where the ST values are. With an if statement it will look if the ST values are in the rows or columns.
          if row is not equal to zero then the ST values are in the rows, so now it needs to skip the row with the ST values
          it does that by checking if the counter is between non_mod_skip - 12 and non_mod_skip because if its in bewteen
          that means that is the row with the ST values. For the columns it checks if two positions are the same as the counter
          check_first_col and check_second_col will be + 12 to move down a row so all columns will be checked.
          Now the function will check if the analysis has mod/non-mods if it doesn't have them then its sets non_mod_limit
          on 12 otherwise it will different between 6 for rows and 7 for columns. now just like with the ST values it will
          check two things for rows it will check if the counter is between and for columns if it is the same.
          After doing all that it will fill up the final_dictionary and final_list by checking
          if they pass any off the requirement given by the if-statement. If all the requirements are met then the list
          is given an 1, if they are not met then the list gets an 0. After the list is filled and sorted
          the list is given to the render.
    """
    try:
        global final_list
        global cut_off_value_au
        global final_dictionary
        global end_result
        global rule
        OD_multiplier = 'None'
        OD_multiplier2 = 'nothing'
        if request.method == 'POST':
            if request.POST.get('download'):
                file_name = request.POST.get('File_name')
                textfile = open("../Download_files/" + file_name + ".txt", "w")
                for elements in final_list:
                    for element in elements:
                        textfile.write(str(element) + "\t")
                    textfile.write("\n")
                textfile.close()
            if request.POST.get('update_table_M') or request.POST.get('update_table_H') or\
                    request.POST.get('update_table_S') or request.POST.get('update_table_No'):
                final_dictionary = {}
                OD_multiplier = request.POST.get('OD_multiplier')
                first_value = list(end_result.values())[0]
                if len(first_value[0]) == 2:
                    for keys, values in dictionary.items():
                        if keys == HD:
                            params = params_dictionary[HD]
                            cut_off_value_au = formula2(float(cut_off_value), *params) * 1
                        elif HD == 'None':
                            cut_off_value_au = 0
                        if keys not in delete:
                            check_first_col = int(column_standard[0]) - 1
                            check_second_col = int(column_standard[1]) - 1
                            counter = 0
                            row = values[1:]
                            if int(row_standard) != 0:
                                non_mod_skip = 12 * int(row_standard)
                            for OD_list in row:
                                well = OD_list[0][0]
                                plate_number = 1
                                column = OD_list[1:]
                                for OD in column:
                                    if int(row_standard) != 0:
                                        if counter < (non_mod_skip - 12) or counter >= non_mod_skip:
                                            if int(OD[0]) > 5:
                                                end_result[keys][counter].append("( + )")
                                            else:
                                                end_result[keys][counter].append(OD[0])
                                            end_result[keys][counter].append(well)
                                            end_result[keys][counter].append(plate_number)
                                        counter += 1
                                        plate_number += 1
                                    else:
                                        if counter == check_second_col:
                                            check_first_col += 12
                                            check_second_col += 12
                                        elif counter != check_first_col:
                                            if int(OD[0]) > 5:
                                                end_result[keys][counter].append("( + )")
                                            else:
                                                end_result[keys][counter].append(OD[0])
                                            end_result[keys][counter].append(well)
                                            end_result[keys][counter].append(plate_number)
                                        counter += 1
                                        plate_number += 1
                sampleID = 1
                final_list = []
                for keys, values in end_result.items():
                    if keys != HD:
                        counter = 0
                        counter2 = 0
                        mod_check_colum1 = int(column_standard[0]) - 1
                        mod_check_colum2 = int(column_standard[1]) - 1
                        for elements in values:
                            if int(row_standard) != 0:
                                non_mod_count = 6
                                if elisa_type == '1':
                                    non_mod_limit = 6
                                else:
                                    non_mod_limit = 12
                                non_mod_skip = 12 * int(row_standard)
                                if counter2 < (non_mod_skip - 12) or counter2 >= non_mod_skip:
                                    non_mod_check = True
                                else:
                                    non_mod_check = False
                            elif int(row_standard) == 0:
                                non_mod_count = 5
                                if elisa_type == '1':
                                    non_mod_limit = 7
                                else:
                                    non_mod_limit = 12
                                if counter2 == mod_check_colum1:
                                    non_mod_check = False
                                elif counter2 == mod_check_colum2:
                                    mod_check_colum1 += 12
                                    mod_check_colum2 += 12
                                    non_mod_check = False
                                else:
                                    non_mod_check = True
                            if elements[0] != 'Empty':
                                if counter < non_mod_limit:
                                    if non_mod_check:
                                        if float(elements[1]) >= float(lower):
                                            if float(elements[1]) <= float(upper):
                                                if elements[1] >= float(cut_off_value_au):
                                                    if elisa_type == '1':
                                                        end_variable = [keys, values[counter2][4],
                                                                        values[counter2][3], str(elements[0]), 1,
                                                                        round(elements[1]), values[counter2][2],
                                                                        values[counter2 + non_mod_count][2]]
                                                    else:
                                                        end_variable = [keys, values[counter2][4],
                                                                        values[counter2][3], str(elements[0]), 1,
                                                                        round(elements[1]), values[counter2][2]]
                                                    if request.POST.get('update_table_M'):
                                                        rule = 1
                                                        OD_multiplier = request.POST.get('OD_multiplier')
                                                        if (values[counter2][2])/(values[counter2 + non_mod_count][2]) >= int(OD_multiplier):
                                                            final_dictionary[sampleID] = end_variable
                                                    elif request.POST.get('update_table_H'):
                                                        rule = 2
                                                        OD_multiplier = request.POST.get('OD_higher')
                                                        if (values[counter2][2]) - (values[counter2 + non_mod_count][2]) >= int(OD_multiplier):
                                                            final_dictionary[sampleID] = end_variable
                                                    elif request.POST.get('update_table_No'):
                                                        rule = 4
                                                        final_dictionary[sampleID] = end_variable
                                                    elif request.POST.get('update_table_S'):
                                                        rule = 3
                                                        OD_multiplier = request.POST.get('OD_multiplier')
                                                        OD_multiplier2 = request.POST.get('reference')
                                                        if OD_multiplier == '':
                                                            OD_multiplier = request.POST.get('OD_higher')
                                                            if OD_multiplier != '':
                                                                rule = '2 and 3'
                                                                if (values[counter2][2]) - (
                                                                values[counter2 + non_mod_count][2]) >= int(OD_multiplier):
                                                                    if (round(elements[1])) >= int(OD_multiplier2):
                                                                        final_dictionary[sampleID] = end_variable
                                                        elif OD_multiplier != None:
                                                            rule = '1 and 3'
                                                            if (values[counter2][2]) / (
                                                            values[counter2 + non_mod_count][2]) >= int(OD_multiplier):
                                                                if (round(elements[1])) >= int(OD_multiplier2):
                                                                    final_dictionary[sampleID] = end_variable
                                                        else:
                                                            if (round(elements[1])) >= int(OD_multiplier2):
                                                                final_dictionary[sampleID] = end_variable
                                        if sampleID not in final_dictionary:
                                            if float(elements[1]) < float(lower):
                                                if elisa_type == '1':
                                                    final_dictionary[sampleID] = [keys, values[counter2][4], values[counter2][3],
                                                                              str(elements[0]), 0, '<' + str(lower),
                                                                              values[counter2][2], values[counter2 + non_mod_count][2]]
                                                else:
                                                    final_dictionary[sampleID] = [keys, values[counter2][4],
                                                                                  values[counter2][3],
                                                                                  str(elements[0]), 0, '<' + str(lower),
                                                                                  values[counter2][2]]
                                            elif float(elements[1]) >= float(upper):
                                                if elisa_type == '1':
                                                    final_dictionary[sampleID] = [keys, values[counter2][4], values[counter2][3],
                                                                              str(elements[0]), 1, '>' + str(upper),
                                                                              values[counter2][2], values[counter2 + non_mod_count][2]]
                                                else:
                                                    final_dictionary[sampleID] = [keys, values[counter2][4],
                                                                                  values[counter2][3],
                                                                                  str(elements[0]), 1, '>' + str(upper),
                                                                                  values[counter2][2]]
                                            else:
                                                if elisa_type == '1':
                                                    final_dictionary[sampleID] = [keys, values[counter2][4], values[counter2][3],
                                                                                  str(elements[0]), 0, round(float(elements[1])),
                                                                                  values[counter2][2], values[counter2 + non_mod_count][2]]
                                                else:
                                                    final_dictionary[sampleID] = [keys, values[counter2][4],
                                                                                  values[counter2][3],
                                                                                  str(elements[0]), 0,
                                                                                  round(float(elements[1])),
                                                                                  values[counter2][2]]
                                sampleID += 1
                            counter += 1
                            counter2 += 1
                            if counter == 12:
                                counter = 0
                for i, lists in final_dictionary.items():
                    final_list.append(lists)
                final_list = sorted(final_list, key=itemgetter(3))
        return render(request, 'End_results.html', {
            'final_list': final_list,
            'upper': upper,
            'lower': lower,
            'cut_off_value': round(cut_off_value_au),
            'rule': rule,
            'rule_value' : OD_multiplier,
            'rule_value2' : OD_multiplier2,
            'unit': unit_name,
            'elisa_type': elisa_type,
            'cut_off_type': cut_off_type,
        })
    except:
        return render(request, 'Error.html', {
         'error': 'An error occurred, please make sure you have submitted all the settings on previous pages.'
        })


def session_writeout(session_name):  # Note: currently used pickle version = 4, supported from py 3.4 and default from py 3.8
    session_name += ".ELISA_App"
    with open(session_name, 'wb') as f:
        pickle.dump((totaal, check, check2, dilution, seprate_dilution, end_dilution, dictionary, HD, delete,
                     points_dictionary, mean_ST_dictionary, mean,
                     std, mean2, std2, check_cut_off, cut_data, outlier_value, cut_off_value, end_result, lower, upper,
                     intermediate_dictionary, params_dictionary, final_dictionary, final_list, cut_off_value_au,
                     unit_name, row_standard, column_standard, elisa_type, cut_off_type,
                     serializers.serialize("xml", Plates.objects.all())), f, protocol=4)  # Plates.objects is serialized to xml, preventing upgrading issues with Django
        print("pickle success")
        f.close()


def session_readin(session):
    varlist = (
        "totaal", "check", "check2", "dilution", "seprate_dilution", "end_dilution", "dictionary", "HD", "delete",
        "points_dictionary", "mean_ST_dictionary", "mean",
        "std", "mean2", "std2", "check_cut_off", "cut_data", "outlier_value", "cut_off_value", "end_result", "lower",
        "upper", "intermediate_dictionary", "params_dictionary", "final_dictionary", "final_list", "cut_off_value_au",
        "unit_name", "row_standard", "column_standard", "elisa_type", "cut_off_type")

    with session as f:
        sessiontuple = pickle.load(f)
        # [:-1] to exclude serialized plate db
        for data, var in zip(sessiontuple[:-1], varlist):
            globals()[var] = data
            print(var)
            print(data, end="\n\n")
        # start plate db readin
        Plates.objects.all().delete()
        for plate in serializers.deserialize("xml", sessiontuple[-1]):
            plate.save()


def autosave(minutes_between_saves = 5): #path here is the directory path, Path refers to the resolve lib, should probably rename the import?
    global last_autosave
    time = datetime.datetime.now()
    if (time - last_autosave).seconds / 60 >= minutes_between_saves:
        last_autosave = time
        path = join(Path(settings.BASE_DIR).resolve().parent, "Autosaves")
        dircontents = listdir(path)
        session_writeout(time.strftime(join("Autosaves", "Autosave %d-%m-%Y  %H.%M.%S")))
        if len(dircontents) > 5:
            remove(min([join(path, session) for session in dircontents], key=getctime)) #Get the oldest file in the dir and remove it


def report_writeout():
    #todo stuff for better report function visualisation, not finished
    """
    global flow
    print([_.id for _ in Plates.objects.all()]) # Inserted Plates Names
    for i in flow["ModifiedLayout"]:    #modified plates
        for j in i:
            for k in j:
                print(k, "\t", sep='', end='')
            print(end="\n")
    for i in flow["PureLayout"]:    #unmodified plates
        for j in i:
            print(j, "\t", sep='', end='')
        print(end="\n")
    for i in dilution: #dilution table
        for j in i:
            for k in j:
                print(k, "\t", sep='', end='')
            print(end="\n")
    for key, value in dictionary.items():
        print(key)
        for i in value:  # modified plates
            for j in i:
                for k in j:
                    print(k, "\t", sep='', end='')
                print(end="\n")
    print(HD) #Selected Healthy Donor
    for i in delete: #Plates to be excluded
        print(i)

    for key, value in flow.items():
        print(key)
        print(value)
    """
    #Create directory, checking for uniqueness
    dirpath = join("Reports", datetime.datetime.now().strftime("Report %d-%m-%Y  %H.%M"))
    unique, iterations = False, 1
    while not unique:
        try:
            mkdir(dirpath)
            unique = True
        except FileExistsError:
            print("FileExistsError raised")
            iterations += 1
            dirpath = (dirpath.split(" (")[0] + f" ({iterations})")

    #Save images
    for file in listdir(get_mediapath()):
        if file.endswith('.png'):
            shutil.copy2(get_mediapath(file), dirpath)

    #Save end results
    with open(join(dirpath, "end_results.txt"), "w") as f:
        f.write(f"Plate name\t"
                f"Plate number\t"
                f"Well number\t"
                f"Sample ID\t"
                f"Positive (1) or Negative (0)\t"
                f"{unit_name}\t"
                f"OD of mod-peptide\t"
                f"OD of non-mod-peptide\n")
        for elements in final_list:
            for element in elements:
                f.write(str(element) + "\t")
            f.write("\n")
        f.close()
    return 0
